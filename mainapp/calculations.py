import math
from mainapp.models import Calibration,Calibration2,Tank_calibration_excel,Fraction
from openpyxl import Workbook,load_workbook
from decimal import Decimal



def cel_to_far(temp):
    temp_int=round(float(temp))
    return round((temp_int*9/5)+32)

#delta 
delta_60= 0.01374979547

# convert temperature f90 to f68
def temp_f_90_to_68(temp_f_90):
    temp_c_90 = (temp_f_90-32)/1.8
    tow = temp_c_90/630
    delta_t = (-0.148759+(-0.267408+(1.080760+(1.269056+(-4.089591+(-1.871251+(7.438081+(-3.536296)*tow)*tow)*tow)*tow)*tow)*tow)*tow)*tow
    temp_c_68=temp_c_90-delta_t
    temp_f_68= 1.8*temp_c_68 + 32
    return temp_f_68

def convert_api_to_dens(api):
    rou= (141.5*999.016)/(api+131.5)
    return rou

def commodity(denss,pro):
    dens = float(denss)
    if(pro==0):
        k0=341.0957
        k1=0
        k2=0
    if(pro==1):
        if(dens>= 838.3127 and dens<=1163.5):
            k0=103.8720
            k1=0.2701
            k2=0
        elif(dens>=787.5195 and dens<838.3127):
            k0=330.3010
            k1=0
            k2=0
        elif(dens>=770.3520 and 787.5195):
            k0=1489.0670
            k1=0
            k2=-0.00186840
        elif(dens>=610.6 and dens<770.3520):
            k0=192.4571
            k1= 0.2438
            k2=0
    return k0,k1,k2

#calculation of correction factor for steel expansion
def correction_expansion_steel(temp_amb,temp_liq):
    t_stl=(temp_amb +(7*temp_liq))/8
    dts= t_stl-60
    a=0.0000062
    cts= 1+(2*dts*a)+(dts*dts)*(a*a)
    return cts

def natural_to_60degree_litr(media_temp,env_temp,natural_litr,density,media_type=1) :

# t_obs_str = input("please enter temperature in f \n ")
    t_obs= float(media_temp)
# product_or_crude_str= input("0 crude /1 product \n")
    product_or_crude= int(media_type)
# density_obs_str = input("enter API \n")
    density_obs=float(density)
# v_obs_str= input("enter Volume \n")
    v_obs= natural_litr
# t_amb_str= input("enter ambient temperature \n")
    t_amb= float(env_temp)
# t_liq_str= input("enter liquid temperature \n")
    t_liq= t_obs

    t_68= temp_f_90_to_68(t_obs)
    # rou_60 = convert_api_to_dens(density_obs)
    rou_60 = density_obs
    k_constand = commodity(rou_60,product_or_crude)

    a = (delta_60/2)*((((k_constand[0]/rou_60)+k_constand[1])/rou_60)+k_constand[2])

    b= (2*k_constand[0]+k_constand[1]*rou_60)/(k_constand[0]+(k_constand[1]+k_constand[2]*rou_60)*rou_60)

    rou_star= rou_60*(1+(math.exp(a*(1+0.8*a))-1)/(1+a*(1+1.6*a)*b))

    alfa_60= (((k_constand[0]/rou_star)+k_constand[1])/rou_star)+k_constand[2]

    delta_t= t_68-60.0068749

    ctl_factor= math.exp(-alfa_60*delta_t*(1+0.8*alfa_60*(delta_t+delta_60)))

    fp_factor= math.exp(-1.9947+0.00013427*t_68+((793920+2326.0*t_68)/(rou_star*rou_star)))

    # cpl_factor=1/(1-0.000001*fp_factor*p_obs) 
    # cptl_factor= cpl_factor*ctl_factor
    # rou=ctl_factor*cpl_factor*rou_star
    cts_factor= correction_expansion_steel(t_amb,t_liq)
    v_60=v_obs*ctl_factor*cts_factor
    # print("volum 60=" + str(v_60))
    # print("ctl 60=" + str(ctl_factor))
    # print("cpl= " + str(cpl_factor))
    # print("cptl = "+ str(cptl_factor))
    # print("rou= "+ str(rou))
    # print("alfa 60 = " + str(alfa_60))
    # print("cts= "+ str(cts_factor))
    return int(v_60)

def litr_excel_MM(id,size):
    tank=Tank_calibration_excel.objects.get(tank_id=id)
    wb=load_workbook(filename=tank.Calibration_excel)
    wb.active
    mm=int(size)
    sheet =wb['Sheet1']
    for row in sheet.iter_rows(0):            
        if row[0].value==mm:
            litr=row[1].value
            break
    return int(litr)
    
def litr_excel_CM(id,size):
    tank=Tank_calibration_excel.objects.get(tank_id=id)
    wb=load_workbook(filename=tank.Calibration_excel)
    wb.active
    cm=int(int(size)/10)
    mm=int(int(size)%10)
    litr_cm=0
    litr_mm=0
    sheet_cm =wb['Sheet1']
    for row in sheet_cm.iter_rows(0):            
        if row[0].value==cm:
            litr_cm=row[1].value
            break
    sheet_mm =wb['Sheet2']
    for row in sheet_mm.iter_rows(0):            
        if row[0].value==mm:
            litr_mm=row[1].value
            break
        else:
            litr_mm=0
    total_litr=int(litr_cm+litr_mm)
    return total_litr

def litr_calibration_old_db(id,size):
    size_previous=Calibration.objects.filter(tankid=id,size__lt=size).last()
    fraction_mm= int(size) - int(size_previous.size)
    fraction_mm_litr=Fraction.objects.get(tankid=id,milno=Decimal(fraction_mm)).litr
    nat_lit=int(size_previous.liters + fraction_mm_litr)
    return nat_lit

def litr_calibration2_old_db(id,size):
    size_previous=Calibration2.objects.filter(tankid=id,size__lt=size).last()
    fraction_mm= int(size) - int(size_previous.size)
    fraction_mm_litr=Fraction.objects.get(tankid=id,milno=Decimal(fraction_mm)).litr
    nat_lit=int(size_previous.liters + fraction_mm_litr)
    return nat_lit

def natural_litr(tank_id,tank_size,water):
    if Tank_calibration_excel.objects.filter(tank_id=tank_id).exists():  # Tank with excel calibration
        tank=Tank_calibration_excel.objects.get(tank_id=tank_id)
        id=tank_id
        size=tank_size
        if tank.Calibration_in_milimeter == True:     
            litr=litr_excel_MM(id,size)-litr_excel_MM(id,water)
            return litr
        else:
            if water != 0: 
                litr=litr_excel_CM(id,size)-litr_excel_CM(id,water)
            else:
                litr=litr=litr_excel_CM(id,size)
            return litr
    else:                                                                      
        if Calibration.objects.filter(tankid=tank_id).exists():                         #Tank without Excel calibration
            if Calibration.objects.filter(tankid=tank_id,size=tank_size).exists():
                nat_lit= Calibration.objects.get(tankid=tank_id,size=tank_size).liters
                if water != 0 :
                    if Calibration.objects.filter(tankid=tank_id,size=water).exists():
                        water_lit= Calibration.objects.get(tankid=tank_id,size=water).liters
                    else:
                        water_lit=litr_calibration_old_db(tank_id,water)
                    nat_lit= nat_lit-water_lit                
                    return int(nat_lit)
                else:
                    water_lit=0
                    return(nat_lit)

            else:
                # print(tank_id,tank_size,water)
                nat_lit=litr_calibration_old_db(tank_id,tank_size)
                if water != 0 :
                    if Calibration.objects.filter(tankid=tank_id,size=water).exists():
                        water_lit= Calibration.objects.get(tankid=tank_id,size=water).liters
                    else:
                        water_lit=litr_calibration_old_db(tank_id,water)
                else:
                    water_lit=0
                x= int(nat_lit)-int(water_lit)
                return x

        else:
            if Calibration2.objects.filter(tankid=tank_id).exists():
                if Calibration2.objects.filter(tankid=tank_id,size=tank_size).exists():
                    nat_lit= Calibration2.objects.get(tankid=tank_id,size=tank_size).liters
                    if water != 0 :
                        if Calibration2.objects.filter(tankid=tank_id,size=water).exist():
                            water_lit= Calibration2.objects.get(tankid=tank_id,size=water).liters
                        else:
                            water_lit=litr_calibration2_old_db(tank_id,water)
                        nat_lit= nat_lit-water_lit                
                        return int(nat_lit)
                else:
                    nat_lit=litr_calibration2_old_db(tank_id,tank_size) - litr_calibration2_old_db(tank_id,water)
                    return nat_lit
            else:
                nat_lit =0
                return nat_lit

        #     nat_lit= Calibration2.objects.get(tankid=tank_id,size=tank_size).liters
        #     if water != 0 :
        #         water_lit= Calibration2.objects.get(tankid=tank_id,size=water).liters
        #         nat_lit= nat_lit-water_lit
        #     return int(nat_lit)

